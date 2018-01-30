import optparse
import ipaddress
from openpyxl import load_workbook
import re

parser = optparse.OptionParser(usage="Usage: %prog [OPTION]... <EXCEl_SPREADSHEET>...", description="Verify overlap of each IPv4 network in a spreadsheet column of multiple spreadsheets.\nIP-address or full CIDR syntax is needed: x.x.x.x/y, only one per cell.",
	epilog='Open Source MIT License. Written by Christian Angerbjorn')
parser.add_option("-n", "--network", default='A', help="Network data column")
parser.add_option("-c", "--comment", default='D', help="Uniqueness comment column")
parser.add_option("-l", "--location", default='B', help="Location column")
parser.add_option("-p", "--location-prepend", help="Prepend this string to all cells that does not already contain this, for example: --location-prepend France")
parser.add_option("-C", "--clean", action="store_true", help="Remove all overlap comment and exit")
parser.add_option("-s", "--stats",  action="store_true", help="Show stats on how many unique IP addresses found.")
parser.add_option("-r", "--read-only",  action="store_true", default=False, help="Open spreadsheets as read-only")
(ops, args) = parser.parse_args()

if len(args) == 0:
	parser.error("At least one excel .xlsx file is required!")

hashTag = "# Overlap:"
excelTag = "Overlap with net"

nets = []  # list of ipaddress.ip_network(  )
directory = {}  # map of ipaddress.ip_network() : 'excel:cell'
for excel in args:
	wb = load_workbook( filename = excel, read_only=ops.read_only)
	for sheet in wb:
		for i, row in enumerate(sheet.iter_rows()):
			# remove any hashTag comments first.
			if not ops.read_only and sheet[ ops.network + str(i+1) ].value and sheet[ ops.network + str(i+1) ].value.startswith(hashTag):
				sheet[ ops.network + str(i+1) ].value = sheet[ ops.network + str(i+1) ].value[len(hashTag) : ] # remove # overlap at begening of string
			if  ops.clean:
				if not ops.read_only and sheet[ ops.comment + str(i+1) ].value and sheet[ ops.comment + str(i+1) ].value.startswith(excelTag):
					sheet[ ops.comment + str(i+1) ].value = ''
			else:
				data = row[ ord(ops.network.upper())-65 ].value
				this_address = '%s:%s%d' %(excel, ops.network.upper(), i+1 )
				try:
					n = ipaddress.ip_network( data )
					# prepend location data 
					if ops.location_prepend:
						loc = row[ ord(ops.location.upper())-65 ].value 
						if (not loc ) or (loc and loc.lower().find( ops.location_prepend.lower() ) == -1):
							row[ ord(ops.location.upper())-65 ].value = "%s - %s" %(ops.location_prepend, loc)
				except ValueError as err:
					if i and not data.startswith('#'): # first line assumed to be a header...
						print( "Warning:    Cell %s%d value '%s' is not recognised as a IPv4 net." %(ops.network, i+1, data))
					continue
	
				# check if this net overlap any existing:
				overlap = False
				for net in nets:
					if net.overlaps( n ):
						overlap = True
						# check not overwrting data:
						if sheet[ ops.comment + str(i+1) ].value and not sheet[ ops.comment + str(i+1) ].value.startswith(excelTag):
							print("Warning: existing data in cell '%s', will not overwrite! %s overlap with net %s cell %s" %(this_address, data, net, directory[ net ]))
						else:
							if not ops.read_only:
								sheet[ ops.comment + str(i+1) ] = "%s %s from %s" %(excelTag, net, directory[ net ])
								sheet[ ops.network + str(i+1) ] = hashTag + sheet[ ops.network + str(i+1) ].value 
							print("  %s -> %s  %s -> %s" %( data, net, this_address, directory[ net ] ))
						break
				# clean previous comment 
				if not overlap and not ops.read_only and sheet[ ops.comment + str(i+1) ].value and sheet[ ops.comment + str(i+1) ].value.startswith(excelTag):
					sheet[ ops.comment + str(i+1) ].value = ''
					
				nets.append( n )
				directory[ n ] = this_address
	if not ops.read_only:
		wb.save(filename = excel)
	
if ops.stats:
	tot = 0
	for net in nets:
		tot += net.num_addresses
	print("Stats: %d unique subnets" %len(nets))
	print("Stats: %d unique addresses" %tot)
	